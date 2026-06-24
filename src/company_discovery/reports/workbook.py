from __future__ import annotations

from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

COMPANY_SHEET = "Companies"
CONTACT_SHEET = "Contacts"
WORKBOOK_FILENAME = "leads.xlsx"

COMPANY_FIELDS = [
    "company_name",
    "domain",
    "bucket",
    "outcome",
    "linkedin_url",
    "phone",
    "street_address",
    "city",
    "state",
    "zip",
    "country",
    "vertical",
    "target_vertical",
    "employee_min",
    "employee_max",
    "ownership_type",
    "independence_status",
    "fit",
    "reason",
    "reason_codes",
    "evidence",
    "source",
    "conflicts",
    "review_flags",
]

CONTACT_FIELDS = [
    "company_name",
    "company_domain",
    "contact_name",
    "title",
    "linkedin_url",
    "email",
    "phone",
    "bucket",
    "status",
    "roles",
    "notes",
    "source",
    "current_company_match",
    "role_match",
    "identity_clear",
    "review_flags",
]


def update_company_discovery_workbook(
    artifacts_root: Path, payload: dict[str, Any]
) -> str:
    run_id = str(payload["run_id"])
    rows = [
        _company_discovery_row(item)
        for item in payload.get("candidates", [])
        if item.get("bucket") in {"selected", "reserve"}
    ]
    path = _workbook_path(artifacts_root, run_id)
    _update_workbook(path, company_rows=rows)
    return str(path.resolve())


def update_company_enrichment_workbook(
    artifacts_root: Path, payload: dict[str, Any]
) -> str | None:
    run_id = str(payload["discovery_run_id"])
    bucket = str(payload.get("bucket") or "")
    rows = [_company_enrichment_row(item, bucket) for item in payload.get("items", [])]
    path = _workbook_path(artifacts_root, run_id)
    merged = _merge_rows(_read_sheet_rows(path, COMPANY_SHEET), rows, _company_key)
    _update_workbook(path, company_rows=merged)
    return str(path.resolve())


def update_contact_discovery_workbook(
    artifacts_root: Path, payload: dict[str, Any]
) -> str:
    run_id = str(payload["source_discovery_run_id"])
    rows = _accepted_contact_rows(payload.get("items", []))
    path = _workbook_path(artifacts_root, run_id)
    merged = _merge_rows(_read_sheet_rows(path, CONTACT_SHEET), rows, _contact_key)
    _update_workbook(path, contact_rows=merged)
    return str(path.resolve())


def update_contact_enrichment_workbook(
    artifacts_root: Path, payload: dict[str, Any]
) -> str:
    run_id = str(payload["source_discovery_run_id"])
    rows = [_contact_enrichment_row(item) for item in payload.get("items", [])]
    path = _workbook_path(artifacts_root, run_id)
    merged = _merge_rows(_read_sheet_rows(path, CONTACT_SHEET), rows, _contact_key)
    _update_workbook(path, contact_rows=merged)
    return str(path.resolve())


def _workbook_path(artifacts_root: Path, discovery_run_id: str) -> Path:
    return artifacts_root / discovery_run_id / WORKBOOK_FILENAME


def _update_workbook(
    path: Path,
    *,
    company_rows: list[dict[str, Any]] | None = None,
    contact_rows: list[dict[str, Any]] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(path) if path.exists() else Workbook()
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        del workbook["Sheet"]
    if company_rows is not None:
        _write_sheet(workbook, COMPANY_SHEET, COMPANY_FIELDS, company_rows, "CompaniesTable", 0)
    elif COMPANY_SHEET not in workbook.sheetnames:
        _write_sheet(workbook, COMPANY_SHEET, COMPANY_FIELDS, [], "CompaniesTable", 0)
    if contact_rows is not None:
        _write_sheet(workbook, CONTACT_SHEET, CONTACT_FIELDS, contact_rows, "ContactsTable", 1)
    elif CONTACT_SHEET not in workbook.sheetnames:
        _write_sheet(workbook, CONTACT_SHEET, CONTACT_FIELDS, [], "ContactsTable", 1)
    workbook.active = workbook[COMPANY_SHEET]
    workbook.save(path)


def _write_sheet(
    workbook: Workbook,
    title: str,
    fields: list[str],
    rows: list[dict[str, Any]],
    table_name: str,
    index: int,
) -> None:
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title, index)
    sheet.append(fields)
    for row in rows:
        sheet.append([_cell_value(row.get(field)) for field in fields])
    if not rows:
        sheet.append(["" for _ in fields])
    _format_sheet(sheet, fields, table_name)


def _read_sheet_rows(path: Path, title: str) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    workbook = load_workbook(path)
    if title not in workbook.sheetnames:
        return []
    sheet = workbook[title]
    headers = [cell.value for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append(
            {
                str(header): value
                for header, value in zip(headers, row, strict=False)
                if header
            }
        )
    return rows


def _merge_rows(
    existing: list[dict[str, Any]],
    updates: list[dict[str, Any]],
    key_fn: Callable[[dict[str, Any]], str],
) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in [*existing, *updates]:
        key = key_fn(row)
        if not key:
            continue
        if key not in merged:
            order.append(key)
            merged[key] = {}
        merged[key] = {**merged[key], **_non_empty(row)}
    return [merged[key] for key in order]


def _format_sheet(sheet: Worksheet, fields: list[str], table_name: str) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    last_column = sheet.cell(row=1, column=len(fields)).column_letter
    last_row = max(sheet.max_row, 2)
    table = Table(displayName=table_name, ref=f"A1:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = min(max(max((len(value) for value in values), default=0) + 2, 12), 60)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _company_discovery_row(item: dict[str, Any]) -> dict[str, Any]:
    company = item.get("company") or {}
    evaluation = item.get("evaluation") or {}
    return {
        "company_name": company.get("company_name"),
        "domain": company.get("domain"),
        "bucket": item.get("bucket"),
        "outcome": item.get("bucket"),
        "state": company.get("state"),
        "country": company.get("country"),
        "vertical": company.get("vertical"),
        "target_vertical": evaluation.get("target_vertical"),
        "employee_min": company.get("employee_min"),
        "employee_max": company.get("employee_max"),
        "ownership_type": company.get("ownership_type"),
        "fit": evaluation.get("fit"),
        "reason": evaluation.get("reason"),
        "reason_codes": _join(evaluation.get("reason_codes")),
        "evidence": _join(evaluation.get("evidence")),
        "source": item.get("source"),
    }


def _company_enrichment_row(item: dict[str, Any], bucket: str) -> dict[str, Any]:
    discovery = item.get("discovery") or {}
    enrichment = item.get("enrichment") or {}
    phone = enrichment.get("phone") or {}
    location = enrichment.get("location") or {}
    independence = enrichment.get("independence") or {}
    linkedin = enrichment.get("linkedin") or {}
    return {
        "company_name": discovery.get("company_name"),
        "domain": discovery.get("domain"),
        "bucket": discovery.get("bucket") or bucket,
        "outcome": item.get("outcome"),
        "linkedin_url": linkedin.get("url"),
        "phone": phone.get("display_value"),
        "street_address": location.get("street_address"),
        "city": location.get("city"),
        "state": location.get("state") or discovery.get("state"),
        "zip": location.get("zip"),
        "country": discovery.get("country"),
        "vertical": discovery.get("vertical"),
        "target_vertical": discovery.get("target_vertical"),
        "employee_min": discovery.get("employee_min"),
        "employee_max": discovery.get("employee_max"),
        "ownership_type": discovery.get("ownership_type"),
        "independence_status": independence.get("status", "unknown"),
        "fit": discovery.get("fit"),
        "reason": discovery.get("reason"),
        "evidence": _join(discovery.get("evidence")),
        "source": discovery.get("source"),
        "conflicts": _join(item.get("conflicts")),
        "review_flags": _join(item.get("review_flags")),
    }


def _accepted_contact_rows(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    verdict_rank = {"rejected": 0, "review": 1, "accepted": 2}
    best_verdict: dict[int, str] = {}
    for item in items:
        candidate_id = int(item["candidate_id"])
        current = best_verdict.get(candidate_id, "rejected")
        if verdict_rank[item["verdict"]] >= verdict_rank[current]:
            best_verdict[candidate_id] = item["verdict"]

    rows: list[dict[str, Any]] = []
    seen: set[int] = set()
    for item in items:
        candidate_id = int(item["candidate_id"])
        best = best_verdict.get(candidate_id)
        if best not in {"accepted", "review"} or candidate_id in seen:
            continue
        seen.add(candidate_id)
        candidate = item.get("candidate") or {}
        related = [
            other
            for other in items
            if int(other["candidate_id"]) == candidate_id and other["verdict"] == best
        ]
        roles = [str(other.get("role_key") or "") for other in related]
        reasons = [str(other.get("reason") or "") for other in related]
        rows.append(
            {
                "company_name": candidate.get("company_name"),
                "company_domain": candidate.get("company_domain"),
                "contact_name": candidate.get("full_name"),
                "title": candidate.get("title"),
                "linkedin_url": candidate.get("linkedin_url"),
                "email": "",
                "phone": "",
                "bucket": best,
                "status": best,
                "roles": _join_unique(roles),
                "notes": _join_unique(reasons),
                "source": item.get("source"),
                "current_company_match": item.get("current_company_match"),
                "role_match": item.get("role_match"),
                "identity_clear": item.get("identity_clear"),
            }
        )
    return rows


def _contact_enrichment_row(item: dict[str, Any]) -> dict[str, Any]:
    discovery = item.get("discovery") or {}
    channels = item.get("channels") or {}
    return {
        "company_name": discovery.get("company_name"),
        "company_domain": discovery.get("company_domain"),
        "contact_name": discovery.get("contact_name"),
        "title": discovery.get("title"),
        "linkedin_url": discovery.get("linkedin_url"),
        "email": channels.get("email"),
        "phone": channels.get("phone"),
        "bucket": "accepted",
        "status": item.get("outcome"),
        "roles": _join(discovery.get("role_keys")),
        "notes": discovery.get("discovery_reason"),
        "review_flags": _join(item.get("review_flags")),
    }


def _cell_value(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, str | int | float | bool):
        return value
    return str(value)


def _non_empty(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if value not in (None, "")}


def _company_key(row: dict[str, Any]) -> str:
    return str(row.get("domain") or "").strip().lower()


def _contact_key(row: dict[str, Any]) -> str:
    domain = str(row.get("company_domain") or "").strip().lower()
    linkedin = str(row.get("linkedin_url") or "").strip().lower()
    name = str(row.get("contact_name") or "").strip().lower()
    title = str(row.get("title") or "").strip().lower()
    return "|".join([domain, linkedin or name, title])


def _join(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list | tuple | set):
        return " | ".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def _join_unique(values: list[str]) -> str:
    return " | ".join(dict.fromkeys(value for value in values if value))
