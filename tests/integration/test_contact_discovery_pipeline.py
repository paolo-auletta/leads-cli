from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from company_discovery.db.contact_repository import ContactDiscoveryRepository
from company_discovery.db.contact_enrichment_repository import ContactEnrichmentRepository
from company_discovery.db.enrichment_repository import EnrichmentRepository
from company_discovery.db.repository import DiscoveryRepository
from company_discovery.domain.contact_models import (
    ApolloBatchResult,
    ApolloPersonMatch,
    ApolloPersonRequest,
    ContactAssessment,
    ContactVerdict,
    EvidenceVerdict,
    ContactSearchBatch,
)
from company_discovery.domain.contact_spec import ContactSearchSpec
from company_discovery.domain.models import (
    CandidateBucket,
    CandidateEvaluation,
    EnrichmentItem,
    EnrichmentOutcome,
    EnrichmentProfile,
    EnrichmentSummary,
    ExaSearchResult,
    ExclusionVerdict,
    FitVerdict,
    MatchVerdict,
    NormalizedCandidate,
    RunSummary,
)
from company_discovery.domain.spec import CompanySearchSpec
from company_discovery.reports.contact_exporter import ContactDiscoveryArtifactExporter
from company_discovery.reports.contact_enrichment_exporter import (
    ContactEnrichmentArtifactExporter,
)
from company_discovery.services.contact_enrichment_pipeline import (
    ContactEnrichmentOptions,
    ContactEnrichmentPipeline,
)
from company_discovery.services.contact_pipeline import ContactDiscoveryPipeline


class FakePeopleSearch:
    def __init__(self) -> None:
        self.calls = 0
        self.last_cost_dollars = 0.01

    def search_people(
        self, query: str, *, country: str, num_results: int
    ) -> list[ExaSearchResult]:
        self.calls += 1
        return [
            ExaSearchResult(
                query=query,
                position=1,
                title="Jane Smith - Project Manager - Acme Builders",
                url="https://www.linkedin.com/in/jane-smith",
                text="Jane Smith is a Project Manager at Acme Builders.",
            )
        ]

    def search_contact_evidence(
        self, query: str, *, country: str, num_results: int
    ) -> list[ExaSearchResult]:
        self.calls += 1
        return [
            ExaSearchResult(
                query=query,
                position=1,
                title="Our Team - Acme Builders",
                url="https://acme.com/team/jane-smith",
                text="Jane Smith is a Project Manager at Acme Builders.",
            )
        ]


class FakeContactEvaluator:
    def evaluate(
        self,
        batch: ContactSearchBatch,
        *,
        current_only: bool,
        require_role_match: bool,
    ) -> list[ContactAssessment]:
        return [
            ContactAssessment(
                full_name="Jane Smith",
                title="Project Manager",
                linkedin_url="https://www.linkedin.com/in/jane-smith",
                source_urls=["https://www.linkedin.com/in/jane-smith"],
                evidence=["Project Manager at Acme Builders"],
                current_company_match=EvidenceVerdict.YES,
                role_match=EvidenceVerdict.YES,
                identity_clear=True,
                verdict=ContactVerdict.ACCEPTED,
                reason="Current company and requested role are explicit.",
            )
        ]


class FakeApollo:
    def __init__(self) -> None:
        self.calls = 0

    def enrich_people(
        self,
        people: list[ApolloPersonRequest],
        *,
        reveal_email: bool,
        reveal_phone: bool,
    ) -> ApolloBatchResult:
        self.calls += 1
        return ApolloBatchResult(
            matches=[
                ApolloPersonMatch(
                    candidate_id=person.candidate_id,
                    person_found=True,
                    full_name=person.full_name,
                    linkedin_url=person.linkedin_url,
                    title="Project Manager",
                    organization_name=person.company_name,
                    organization_domain=person.company_domain,
                    email=f"jane@{person.company_domain}" if reveal_email else None,
                    email_status="verified" if reveal_email else None,
                    phones=["+15125550100"] if reveal_phone else [],
                    apollo_person_id="apollo-jane",
                    raw={"provider": "fake-apollo"},
                )
                for person in people
            ]
        )

    def poll(self, request_id: str) -> ApolloBatchResult:
        raise AssertionError("synchronous fake should not be polled")

    def close(self) -> None:
        pass


def _completed_company_enrichment(repository: DiscoveryRepository) -> str:
    spec = CompanySearchSpec.model_validate(
        {
            "version": 1,
            "count": 1,
            "vertical": {"key": "construction", "label": "Construction"},
        }
    )
    discovery_run_id = repository.create_run(spec)
    candidate_id = repository.upsert_candidate(
        NormalizedCandidate(
            company_name="Acme Builders",
            domain="acme.com",
            dedupe_key="acme.com",
            vertical="construction",
            country="US",
            state="TX",
        )
    )
    repository.record_evaluation(
        discovery_run_id,
        candidate_id,
        CandidateEvaluation(
            company_name="Acme Builders",
            domain="acme.com",
            fit=FitVerdict.GOOD,
            vertical_match=MatchVerdict.YES,
            geography_match=MatchVerdict.YES,
            size_match=MatchVerdict.UNKNOWN,
            excluded=ExclusionVerdict.NO,
            reason="Matches the target",
            reason_codes=[],
            evidence=["Texas builder"],
            inferred_vertical="construction",
            inferred_country="US",
            inferred_state="TX",
            inferred_employee_min=None,
            inferred_employee_max=None,
            inferred_ownership_type="independent",
            target_vertical="construction",
        ),
        CandidateBucket.SELECTED,
        "exa",
    )
    repository.complete_run(discovery_run_id, RunSummary(selected=1), {})

    enrichment_repository = EnrichmentRepository(repository.database)
    enrichment_run_id = enrichment_repository.create_run(
        discovery_run_id, "selected", {}
    )
    enrichment_repository.save_item(
        enrichment_run_id,
        EnrichmentItem(
            company_id=candidate_id,
            discovery={
                "run_id": discovery_run_id,
                "company_name": "Acme Builders",
                "domain": "acme.com",
                "vertical": "construction",
                "target_vertical": "construction",
                "country": "US",
                "state": "TX",
            },
            enrichment=EnrichmentProfile(),
            inherited_status={},
            outcome=EnrichmentOutcome.READY,
        ),
    )
    enrichment_repository.complete_run(
        enrichment_run_id, EnrichmentSummary(processed=1, ready=1), {}
    )
    return enrichment_run_id


def _contact_spec(enrichment_run_id: str) -> ContactSearchSpec:
    return ContactSearchSpec.model_validate(
        {
            "version": 1,
            "company_source": {"enrichment_run_id": enrichment_run_id},
            "roles": [
                {
                    "key": "project_manager",
                    "labels": ["project manager", "senior project manager"],
                    "max_per_company": 1,
                }
            ],
        }
    )


def test_contact_discovery_searches_exports_trace_and_reuses_memory(
    repository: DiscoveryRepository,
    tmp_path: Path,
) -> None:
    enrichment_run_id = _completed_company_enrichment(repository)
    contact_repository = ContactDiscoveryRepository(repository.database)
    search = FakePeopleSearch()
    pipeline = ContactDiscoveryPipeline(
        repository=contact_repository,
        exporter=ContactDiscoveryArtifactExporter(tmp_path / "runs"),
        search_provider=search,
        evaluator=FakeContactEvaluator(),  # type: ignore[arg-type]
        results_per_query=5,
    )

    first = pipeline.discover(_contact_spec(enrichment_run_id))

    assert first.run_id.startswith("contact-discover-")
    assert first.summary.companies_loaded == 1
    assert first.summary.queries_run == 2
    assert first.summary.raw_results == 2
    assert first.summary.accepted == 1
    assert search.calls == 2
    accepted_path = Path(first.artifact_paths["accepted"])
    assert accepted_path.parent.name == first.run_id
    assert accepted_path.parent.parent.name == "contacts"
    assert accepted_path.parent.parent.parent.name == enrichment_run_id
    assert accepted_path.parent.parent.parent.parent.name == "enrich"
    with Path(first.artifact_paths["accepted"]).open() as handle:
        rows = list(csv.DictReader(handle))
    assert list(rows[0]) == ContactDiscoveryArtifactExporter.FIELDS
    assert rows[0] == {
        "company_name": "Acme Builders",
        "company_domain": "acme.com",
        "contact_name": "Jane Smith",
        "title": "Project Manager",
        "linkedin_url": "https://www.linkedin.com/in/jane-smith",
        "email": "",
        "phone": "",
        "status": "accepted",
        "notes": "project_manager: Current company and requested role are explicit.",
    }
    workbook = load_workbook(first.artifact_paths["workbook"])
    contacts = workbook["Contacts"]
    assert contacts["A2"].value == "Acme Builders"
    assert contacts["C2"].value == "Jane Smith"
    assert contacts["F2"].value is None
    assert contacts["G2"].value is None
    assert contacts["H2"].value == "accepted"
    assert contacts["I2"].value == "accepted"
    payload = json.loads(Path(first.artifact_paths["json"]).read_text())
    assert payload["source_discovery_run_id"] == accepted_path.parent.parent.parent.parent.parent.name
    assert len(payload["queries"]) == 2
    assert payload["queries"][0]["raw_results"][0]["text"].startswith("Jane Smith")
    assert payload["items"][0]["candidate"]["evidence"] == [
        "Project Manager at Acme Builders"
    ]

    memory_only = ContactDiscoveryPipeline(
        repository=contact_repository,
        exporter=ContactDiscoveryArtifactExporter(tmp_path / "runs"),
        search_provider=None,
        evaluator=None,
    )
    second = memory_only.discover(_contact_spec(enrichment_run_id))

    assert second.run_id.startswith("contact-discover-")
    assert second.run_id != first.run_id
    assert second.summary.memory_reused == 1
    assert second.summary.queries_run == 0
    assert second.summary.accepted == 1
    assert second.items[0].source == "memory"


def test_contact_spec_rejects_domains_outside_the_selected_enrichment_bucket(
    repository: DiscoveryRepository,
    tmp_path: Path,
) -> None:
    enrichment_run_id = _completed_company_enrichment(repository)
    payload = _contact_spec(enrichment_run_id).model_dump()
    payload["company_source"]["domains"] = ["other.com"]
    spec = ContactSearchSpec.model_validate(payload)
    pipeline = ContactDiscoveryPipeline(
        repository=ContactDiscoveryRepository(repository.database),
        exporter=ContactDiscoveryArtifactExporter(tmp_path / "runs"),
        search_provider=None,
        evaluator=None,
    )

    try:
        pipeline.discover(spec)
    except ValueError as exc:
        assert "other.com" in str(exc)
    else:
        raise AssertionError("out-of-scope domain should fail before contact discovery")


def test_contact_enrichment_uses_accepted_people_exports_under_contact_run_and_reuses_memory(
    repository: DiscoveryRepository,
    tmp_path: Path,
) -> None:
    enrichment_run_id = _completed_company_enrichment(repository)
    contact_repository = ContactDiscoveryRepository(repository.database)
    discovery = ContactDiscoveryPipeline(
        repository=contact_repository,
        exporter=ContactDiscoveryArtifactExporter(tmp_path / "runs"),
        search_provider=FakePeopleSearch(),
        evaluator=FakeContactEvaluator(),  # type: ignore[arg-type]
    ).discover(_contact_spec(enrichment_run_id))

    apollo = FakeApollo()
    enrichment_repository = ContactEnrichmentRepository(repository.database)
    pipeline = ContactEnrichmentPipeline(
        repository=enrichment_repository,
        exporter=ContactEnrichmentArtifactExporter(tmp_path / "runs"),
        provider=apollo,
        poll_interval_seconds=0.01,
    )

    first = pipeline.enrich(discovery.run_id)

    assert first.run_id.startswith("contact-enrich-")
    assert first.summary.contacts_loaded == 1
    assert first.summary.apollo_requests == 1
    assert first.summary.ready == 1
    assert first.items[0].channels.email == "jane@acme.com"
    assert first.items[0].channels.phone == "+15125550100"
    assert first.items[0].discovery["title"] == "Project Manager"
    ready_path = Path(first.artifact_paths["ready"])
    assert ready_path.parent.name == first.run_id
    assert ready_path.parent.parent.name == "enrich"
    assert ready_path.parent.parent.parent.name == discovery.run_id
    assert ready_path.parent.parent.parent.parent.name == "contacts"
    with ready_path.open() as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "company_name": "Acme Builders",
            "company_domain": "acme.com",
            "contact_name": "Jane Smith",
            "title": "Project Manager",
            "linkedin_url": "https://www.linkedin.com/in/jane-smith",
            "email": "jane@acme.com",
            "phone": "+15125550100",
            "status": "ready",
            "notes": "",
        }
    ]
    workbook = load_workbook(first.artifact_paths["workbook"])
    contacts = workbook["Contacts"]
    assert contacts["C2"].value == "Jane Smith"
    assert contacts["F2"].value == "jane@acme.com"
    assert contacts["G2"].value == "+15125550100"
    assert contacts["H2"].value == "accepted"
    assert contacts["I2"].value == "ready"
    run_payload = json.loads(Path(first.artifact_paths["json"]).read_text())
    assert run_payload["items"][0]["trace"][0]["provider_record"] == {
        "provider": "fake-apollo"
    }

    second = pipeline.enrich(
        discovery.run_id,
        options=ContactEnrichmentOptions(reveal_email=True, reveal_phone=True),
    )

    assert second.run_id.startswith("contact-enrich-")
    assert second.run_id != first.run_id
    assert second.summary.memory_reused == 1
    assert second.summary.apollo_requests == 0
    assert second.summary.ready == 1
    assert apollo.calls == 1

    email_only = pipeline.enrich(
        discovery.run_id,
        options=ContactEnrichmentOptions(reveal_email=True, reveal_phone=False),
    )

    assert email_only.summary.memory_reused == 1
    assert email_only.items[0].channels.email == "jane@acme.com"
    assert email_only.items[0].channels.phone is None
    assert apollo.calls == 1


def test_contact_enrichment_holds_stale_company_match_for_review(
    repository: DiscoveryRepository,
    tmp_path: Path,
) -> None:
    enrichment_run_id = _completed_company_enrichment(repository)
    discovery = ContactDiscoveryPipeline(
        repository=ContactDiscoveryRepository(repository.database),
        exporter=ContactDiscoveryArtifactExporter(tmp_path / "runs"),
        search_provider=FakePeopleSearch(),
        evaluator=FakeContactEvaluator(),  # type: ignore[arg-type]
    ).discover(_contact_spec(enrichment_run_id))

    class StaleApollo(FakeApollo):
        def enrich_people(self, people, *, reveal_email, reveal_phone):
            person = people[0]
            return ApolloBatchResult(
                matches=[
                    ApolloPersonMatch(
                        candidate_id=person.candidate_id,
                        person_found=True,
                        full_name=person.full_name,
                        linkedin_url=person.linkedin_url,
                        organization_name="OldCo",
                        organization_domain="oldco.com",
                        email="jane@oldco.com",
                    )
                ]
            )

    result = ContactEnrichmentPipeline(
        repository=ContactEnrichmentRepository(repository.database),
        exporter=ContactEnrichmentArtifactExporter(tmp_path / "runs"),
        provider=StaleApollo(),
    ).enrich(
        discovery.run_id,
        options=ContactEnrichmentOptions(reveal_email=True, reveal_phone=False),
    )

    assert result.summary.review == 1
    assert "apollo_company_mismatch" in result.items[0].review_flags
    assert result.items[0].discovery["company_name"] == "Acme Builders"
