from __future__ import annotations

from openpyxl import load_workbook

from company_discovery.reports.workbook import (
    update_company_discovery_workbook,
    update_company_enrichment_workbook,
    update_contact_discovery_workbook,
    update_contact_enrichment_workbook,
)


def test_company_workbook_keeps_reserve_companies_when_selected_are_enriched(
    tmp_path,
) -> None:
    root = tmp_path / "runs"
    discovery_payload = {
        "run_id": "company-discover-test",
        "candidates": [
            {
                "bucket": "selected",
                "source": "exa",
                "company": {
                    "company_name": "Acme Builders",
                    "domain": "acme.com",
                    "state": "TX",
                },
                "evaluation": {
                    "fit": "good",
                    "reason": "Matches ICP.",
                    "reason_codes": ["vertical_fit"],
                    "evidence": ["Construction firm"],
                },
            },
            {
                "bucket": "reserve",
                "source": "exa",
                "company": {
                    "company_name": "Beta Builders",
                    "domain": "beta.com",
                    "state": "TX",
                },
                "evaluation": {
                    "fit": "good",
                    "reason": "Reserve fit.",
                    "reason_codes": ["reserve_fit"],
                    "evidence": ["Construction firm"],
                },
            },
        ],
    }
    path = update_company_discovery_workbook(root, discovery_payload)

    enrichment_payload = {
        "discovery_run_id": "company-discover-test",
        "bucket": "selected",
        "items": [
            {
                "discovery": {
                    "company_name": "Acme Builders",
                    "domain": "acme.com",
                    "state": "TX",
                    "fit": "good",
                    "reason": "Matches ICP.",
                    "source": "exa",
                },
                "enrichment": {
                    "phone": {"display_value": "(210) 555-1234"},
                    "linkedin": {"url": "https://www.linkedin.com/company/acme-builders"},
                    "independence": {"status": "yes"},
                },
                "outcome": "enriched_ready",
                "conflicts": [],
                "review_flags": [],
            }
        ],
    }
    update_company_enrichment_workbook(root, enrichment_payload)

    sheet = load_workbook(path)["Companies"]
    assert sheet["A2"].value == "Acme Builders"
    assert sheet["C2"].value == "selected"
    assert sheet["D2"].value == "enriched_ready"
    assert sheet["F2"].value == "(210) 555-1234"
    assert sheet["A3"].value == "Beta Builders"
    assert sheet["C3"].value == "reserve"
    assert sheet["D3"].value == "reserve"


def test_contact_workbook_keeps_review_contacts_when_enrichment_updates_accepted(
    tmp_path,
) -> None:
    root = tmp_path / "runs"
    discovery_payload = {
        "source_discovery_run_id": "company-discover-test",
        "items": [
            {
                "candidate_id": 1,
                "candidate": {
                    "company_name": "Acme Builders",
                    "company_domain": "acme.com",
                    "full_name": "Jane Smith",
                    "title": "Project Manager",
                    "linkedin_url": "https://www.linkedin.com/in/jane-smith",
                },
                "verdict": "accepted",
                "role_key": "project_manager",
                "reason": "Current role confirmed.",
                "source": "exa",
                "current_company_match": "yes",
                "role_match": "yes",
                "identity_clear": True,
            },
            {
                "candidate_id": 2,
                "candidate": {
                    "company_name": "Acme Builders",
                    "company_domain": "acme.com",
                    "full_name": "Sam Taylor",
                    "title": "Operations Manager",
                    "linkedin_url": "https://www.linkedin.com/in/sam-taylor",
                },
                "verdict": "review",
                "role_key": "operations",
                "reason": "Likely current company but evidence is thin.",
                "source": "exa",
                "current_company_match": "likely",
                "role_match": "likely",
                "identity_clear": True,
            },
        ],
    }
    path = update_contact_discovery_workbook(root, discovery_payload)

    enrichment_payload = {
        "source_discovery_run_id": "company-discover-test",
        "items": [
            {
                "discovery": {
                    "company_name": "Acme Builders",
                    "company_domain": "acme.com",
                    "contact_name": "Jane Smith",
                    "title": "Project Manager",
                    "linkedin_url": "https://www.linkedin.com/in/jane-smith",
                    "role_keys": ["project_manager"],
                    "discovery_reason": "Current role confirmed.",
                },
                "channels": {"email": "jane@acme.com", "phone": "+15125550100"},
                "outcome": "ready",
                "review_flags": [],
            }
        ],
    }
    update_contact_enrichment_workbook(root, enrichment_payload)

    sheet = load_workbook(path)["Contacts"]
    assert sheet["C2"].value == "Jane Smith"
    assert sheet["F2"].value == "jane@acme.com"
    assert sheet["I2"].value == "ready"
    assert sheet["C3"].value == "Sam Taylor"
    assert sheet["H3"].value == "review"
    assert sheet["I3"].value == "review"
